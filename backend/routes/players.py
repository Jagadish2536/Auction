import io
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required
from models.database import db
from models.player import Player
from models.tournament import Tournament
from middleware import role_required, tournament_scope_required
from utils import save_upload, allowed_file, clear_tournament_caches
import re

players_bp = Blueprint('players', __name__, url_prefix='/api/tournaments/<int:tournament_id>/players')


def extract_url(text):
    """Extract actual URL from CricHeroes share text like 'Hey! Here is my Cricket profile... https://chshare.link/player/xxx'"""
    if not text:
        return text
    text = text.strip()
    match = re.search(r'https?://[^\s]+', text)
    if match:
        return match.group(0)
    # If no http URL found, check if it looks like a bare domain (e.g. chshare.link/player/xxx)
    if '.' in text and ' ' not in text:
        return f'https://{text}'
    return text


@players_bp.route('', methods=['GET'])
@jwt_required()
@tournament_scope_required
def get_players(tournament_id):
    """Get all players for a tournament with optional pagination."""
    from sqlalchemy.orm import joinedload
    status = request.args.get('status')
    page = request.args.get('page', type=int)
    per_page = request.args.get('per_page', 1000, type=int)

    query = Player.query.filter_by(tournament_id=tournament_id).options(
        joinedload(Player.sold_team)
    )
    if status:
        query = query.filter_by(status=status)
    else:
        query = query.filter(Player.status != 'pending')
    query = query.order_by(Player.name)

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        return jsonify({
            'players': [p.to_dict() for p in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        }), 200

    players = query.all()
    return jsonify({'players': [p.to_dict() for p in players]}), 200


@players_bp.route('/<int:player_id>', methods=['GET'])
@jwt_required()
@tournament_scope_required
def get_player(tournament_id, player_id):
    """Get a single player."""
    player = Player.query.filter_by(id=player_id, tournament_id=tournament_id).first_or_404()
    return jsonify({'player': player.to_dict()}), 200


@players_bp.route('', methods=['POST'])
@role_required('manager', 'admin')
@tournament_scope_required
def create_player(tournament_id):
    """Create a new player."""
    tournament = Tournament.query.get_or_404(tournament_id)
    data = request.form if request.form else request.get_json()

    name = data.get('name')
    if not name:
        return jsonify({'error': 'Player name is required'}), 400

    village = data.get('village')
    mobile_str = str(data.get('mobile') or '').strip()
    playing_style = data.get('playing_style')
    age = data.get('age')

    if not village or not str(village).strip():
        return jsonify({'error': 'Village is required'}), 400
    if not mobile_str:
        return jsonify({'error': 'Mobile number is required'}), 400
    if not mobile_str.isdigit() or len(mobile_str) != 10:
        return jsonify({'error': 'Mobile number must be exactly 10 digits.'}), 400
    if not playing_style or not str(playing_style).strip():
        return jsonify({'error': 'Playing Style is required'}), 400
    if not age:
        return jsonify({'error': 'Age is required'}), 400

    raw_crickheroes = data.get('crickheroes_url')
    if not raw_crickheroes or not str(raw_crickheroes).strip():
        return jsonify({'error': 'CricHeroes Profile URL is required'}), 400

    raw_base_price = data.get('base_price')
    base_price = float(raw_base_price) if (raw_base_price is not None and str(raw_base_price).strip() != '') else (tournament.default_base_price if tournament.default_base_price is not None else 1000.0)

    if mobile_str:
        existing_player = Player.query.filter_by(tournament_id=tournament_id, mobile=mobile_str).first()
        if existing_player:
            return jsonify({'error': 'This mobile number is already registered.'}), 400

    crickheroes_url = extract_url(data.get('crickheroes_url'))

    player = Player(
        tournament_id=tournament_id,
        name=name,
        village=str(village).strip(),
        mobile=mobile_str if mobile_str else None,
        playing_style=str(playing_style).strip(),
        age=int(age) if age else None,
        base_price=base_price,
        status='available',
        crickheroes_url=crickheroes_url
    )

    if hasattr(request, 'files') and 'photo' in request.files:
        photo_path = save_upload(request.files['photo'], 'players')
        if photo_path:
            player.photo = photo_path

    db.session.add(player)
    db.session.commit()
    clear_tournament_caches(tournament_id)
    from app import socketio
    socketio.emit('player:change', {'action': 'created', 'tournament_id': tournament_id, 'player_id': player.id})
    return jsonify({'player': player.to_dict(), 'message': 'Player created successfully'}), 201


@players_bp.route('/<int:player_id>', methods=['PUT'])
@role_required('manager', 'admin')
@tournament_scope_required
def update_player(tournament_id, player_id):
    """Update a player."""
    player = Player.query.filter_by(id=player_id, tournament_id=tournament_id).first_or_404()
    data = request.form if request.form else request.get_json()

    if data.get('name'):
        player.name = data['name']
    if data.get('village') is not None:
        player.village = data.get('village')
    if data.get('mobile') is not None:
        mob_str = str(data.get('mobile') or '').strip()
        if mob_str:
            existing_player = Player.query.filter(
                Player.tournament_id == tournament_id,
                Player.mobile == mob_str,
                Player.id != player_id
            ).first()
            if existing_player:
                return jsonify({'error': 'This mobile number is already registered.'}), 400
            player.mobile = mob_str
        else:
            player.mobile = None
    if data.get('playing_style') is not None:
        player.playing_style = data.get('playing_style')
    if data.get('age') is not None:
        player.age = int(data['age']) if data['age'] else None
    if data.get('base_price') is not None:
        player.base_price = float(data['base_price'])
    if data.get('crickheroes_url') is not None:
        player.crickheroes_url = extract_url(data.get('crickheroes_url'))
    if data.get('status') is not None:
        player.status = data.get('status')

    if hasattr(request, 'files') and 'photo' in request.files:
        photo_path = save_upload(request.files['photo'], 'players')
        if photo_path:
            player.photo = photo_path

    db.session.commit()
    clear_tournament_caches(tournament_id)
    from app import socketio
    socketio.emit('player:change', {'action': 'updated', 'tournament_id': tournament_id, 'player_id': player.id})
    return jsonify({'player': player.to_dict(), 'message': 'Player updated successfully'}), 200


@players_bp.route('/<int:player_id>', methods=['DELETE'])
@role_required('manager', 'admin')
@tournament_scope_required
def delete_player(tournament_id, player_id):
    """Delete a player."""
    player = Player.query.filter_by(id=player_id, tournament_id=tournament_id).first_or_404()
    db.session.delete(player)
    db.session.commit()
    clear_tournament_caches(tournament_id)
    from app import socketio
    socketio.emit('player:change', {'action': 'deleted', 'tournament_id': tournament_id, 'player_id': player_id})
    return jsonify({'message': 'Player deleted successfully'}), 200


@players_bp.route('/bulk-import', methods=['POST'])
@role_required('manager', 'admin')
@tournament_scope_required
def bulk_import_players(tournament_id):
    """Bulk import players from CSV or Excel using csv/openpyxl."""
    tournament = Tournament.query.get_or_404(tournament_id)

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    players_to_add = []

    try:
        if ext == 'csv':
            import csv
            stream = io.StringIO(file.stream.read().decode("utf-8"), newline=None)
            reader = csv.DictReader(stream)
            # Normalize fieldnames
            if not reader.fieldnames:
                return jsonify({'error': 'CSV has no headers'}), 400
            headers = [c.strip().lower().replace(' ', '_') for c in reader.fieldnames]
            
            for row in reader:
                # Map standardized keys
                row_data = {headers[i]: v for i, (k, v) in enumerate(row.items()) if i < len(headers)}
                name = row_data.get('name', '').strip()
                if not name:
                    continue
                
                players_to_add.append({
                    'name': name,
                    'village': row_data.get('village', '').strip() or None,
                    'mobile': row_data.get('mobile', '').strip() or None,
                    'playing_style': row_data.get('playing_style', '').strip() or None,
                    'age': int(row_data['age']) if row_data.get('age') and str(row_data['age']).isdigit() else None,
                    'base_price': float(row_data['base_price']) if row_data.get('base_price') else None
                })
        elif ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True)
            if not wb.sheetnames:
                return jsonify({'error': 'Excel file has no sheets'}), 400
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return jsonify({'error': 'Excel is empty'}), 400
            
            headers = [str(c).strip().lower().replace(' ', '_') for c in rows[0] if c is not None]
            if 'name' not in headers:
                return jsonify({'error': 'Missing required column: name'}), 400
            
            for r in rows[1:]:
                # Map row cells to headers
                row_data = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                name = str(row_data.get('name') or '').strip()
                if not name:
                    continue
                
                players_to_add.append({
                    'name': name,
                    'village': str(row_data.get('village') or '').strip() or None,
                    'mobile': str(row_data.get('mobile') or '').strip() or None,
                    'playing_style': str(row_data.get('playing_style') or '').strip() or None,
                    'age': int(row_data['age']) if row_data.get('age') and str(row_data['age']).isdigit() else None,
                    'base_price': float(row_data['base_price']) if row_data.get('base_price') else None
                })
        else:
            return jsonify({'error': 'Unsupported file format. Use CSV or Excel.'}), 400

        count = 0
        for p_data in players_to_add:
            p_base_price = p_data['base_price'] if p_data['base_price'] is not None else (tournament.default_base_price if tournament.default_base_price is not None else 1000.0)
            player = Player(
                tournament_id=tournament_id,
                name=p_data['name'],
                village=p_data['village'],
                mobile=p_data['mobile'],
                playing_style=p_data['playing_style'],
                age=p_data['age'],
                base_price=p_base_price,
                status='available'
            )
            db.session.add(player)
            count += 1

        db.session.commit()
        clear_tournament_caches(tournament_id)
        from app import socketio
        socketio.emit('player:change', {'action': 'bulk_imported', 'tournament_id': tournament_id})
        return jsonify({'message': f'{count} players imported successfully', 'count': count}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Import failed: {str(e)}'}), 400
