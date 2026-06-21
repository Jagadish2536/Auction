import io
from flask import Blueprint, jsonify, send_file, request
from flask_jwt_extended import jwt_required
from models.player import Player
from models.team import Team
from models.tournament import Tournament
from middleware import role_required, tournament_scope_required

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


def _get_players_data(tournament_id, status=None):
    query = Player.query.filter_by(tournament_id=tournament_id)
    if status:
        query = query.filter_by(status=status)
    return query.order_by(Player.name).all()


@reports_bp.route('/players/<int:tournament_id>', methods=['GET'])
@jwt_required()
@tournament_scope_required
def export_players(tournament_id):
    """Export players report."""
    fmt = request.args.get('format', 'csv')
    status = request.args.get('status')
    players = _get_players_data(tournament_id, status)
    tournament = Tournament.query.get_or_404(tournament_id)

    if fmt == 'csv':
        return _export_players_csv(players, tournament)
    elif fmt == 'excel':
        return _export_players_excel(players, tournament)
    elif fmt == 'pdf':
        return _export_players_pdf(players, tournament)
    elif fmt == 'json':
        return jsonify({
            'players': [p.to_dict() for p in players],
            'tournament': tournament.to_dict()
        })
    return jsonify({'error': 'Invalid format'}), 400


@reports_bp.route('/teams/<int:tournament_id>', methods=['GET'])
@jwt_required()
@tournament_scope_required
def export_teams(tournament_id):
    """Export teams report."""
    fmt = request.args.get('format', 'csv')
    team_id = request.args.get('team_id')
    if team_id:
        teams = Team.query.filter_by(tournament_id=tournament_id, id=int(team_id)).all()
    else:
        teams = Team.query.filter_by(tournament_id=tournament_id).order_by(Team.name).all()
    tournament = Tournament.query.get_or_404(tournament_id)

    if fmt == 'csv':
        return _export_teams_csv(teams, tournament)
    elif fmt == 'excel':
        return _export_teams_excel(teams, tournament)
    elif fmt == 'pdf':
        return _export_teams_pdf(teams, tournament)
    elif fmt == 'json':
        rem_players_pool = Player.query.filter(
            Player.tournament_id == tournament_id,
            Player.status != 'sold',
            Player.status != 'pending'
        ).count()
        return jsonify({
            'teams': [{
                **t.to_dict(),
                'players': [p.to_dict() for p in t.players.all()]
            } for t in teams],
            'tournament': tournament.to_dict(),
            'rem_players_pool': rem_players_pool
        })
    return jsonify({'error': 'Invalid format'}), 400


@reports_bp.route('/auction/<int:tournament_id>', methods=['GET'])
@jwt_required()
@tournament_scope_required
def export_auction(tournament_id):
    """Export full auction report."""
    fmt = request.args.get('format', 'csv')
    sold_players = Player.query.filter_by(tournament_id=tournament_id, status='sold').order_by(Player.sold_at).all()
    tournament = Tournament.query.get_or_404(tournament_id)

    if fmt == 'csv':
        return _export_auction_csv(sold_players, tournament)
    elif fmt == 'excel':
        return _export_auction_excel(sold_players, tournament)
    elif fmt == 'pdf':
        return _export_auction_pdf(sold_players, tournament)
    elif fmt == 'json':
        return jsonify({
            'sold_players': [p.to_dict() for p in sold_players],
            'tournament': tournament.to_dict()
        })
    return jsonify({'error': 'Invalid format'}), 400


# ---- CSV Exports ----

def _export_players_csv(players, tournament):
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['S.No', 'Name', 'Village', 'Mobile', 'Playing Style', 'Age', 'Base Price', 'Status', 'Sold Team', 'Sold Price'])
    for idx, p in enumerate(players):
        writer.writerow([idx + 1, p.name, p.village, p.mobile, p.playing_style, p.age, p.base_price, p.status,
                         p.sold_team.name if p.sold_team else '', p.sold_price or ''])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8')),
                     mimetype='text/csv', as_attachment=True,
                     download_name=f'{tournament.name}_players.csv')


def _export_teams_csv(teams, tournament):
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Team Name', 'Owner Name', 'Captain Name', 'Total Budget', 'Remaining Budget', 'Player Name', 'Playing Style', 'Base Price', 'Sold Price'])
    for t in teams:
        players = t.players.all()
        if not players:
            writer.writerow([t.name, t.owner_name or '', t.captain_name or '', t.budget, t.remaining_budget, 'No Players', '', '', ''])
        else:
            for p in players:
                writer.writerow([t.name, t.owner_name or '', t.captain_name or '', t.budget, t.remaining_budget, p.name, p.playing_style or '', p.base_price, p.sold_price or ''])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8')),
                     mimetype='text/csv', as_attachment=True,
                     download_name=f'{tournament.name}_teams.csv')


def _export_auction_csv(sold_players, tournament):
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['S.No', 'Player', 'Village', 'Playing Style', 'Base Price', 'Sold Price', 'Sold Team', 'Sold At'])
    for idx, p in enumerate(sold_players):
        writer.writerow([idx + 1, p.name, p.village, p.playing_style, p.base_price, p.sold_price,
                         p.sold_team.name if p.sold_team else '', p.sold_at])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8')),
                     mimetype='text/csv', as_attachment=True,
                     download_name=f'{tournament.name}_auction.csv')


# ---- Excel Exports ----

def _export_players_excel(players, tournament):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Players'
    ws.append(['S.No', 'Name', 'Village', 'Mobile', 'Playing Style', 'Age', 'Base Price', 'Status', 'Sold Team', 'Sold Price'])
    for idx, p in enumerate(players):
        ws.append([idx + 1, p.name, p.village, p.mobile, p.playing_style, p.age, p.base_price, p.status,
                   p.sold_team.name if p.sold_team else '', p.sold_price or ''])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{tournament.name}_players.xlsx')


def _export_teams_excel(teams, tournament):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Teams'

    # Font definitions
    font_title = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_stats = Font(name='Segoe UI', size=9, bold=False, color='1E293B')
    font_header = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
    font_data = Font(name='Segoe UI', size=10)
    font_price = Font(name='Segoe UI', size=10, bold=True, color='0F172A')

    # Alignments
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Fills
    fill_team_title = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    fill_stats = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    fill_subheader = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid') # soft gold
    fill_even_row = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    fill_odd_row = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Borders
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for idx, t in enumerate(teams):
        col_start = (idx * 2) + 1
        col_end = (idx * 2) + 2

        # 1. Team Name Row
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        cell_title = ws.cell(row=1, column=col_start)
        cell_title.value = t.name
        cell_title.font = font_title
        cell_title.fill = fill_team_title
        cell_title.alignment = align_center

        # Empty merged partner cell needs style as well
        ws.cell(row=1, column=col_end).fill = fill_team_title

        players = t.players.all()
        purchased = len(players)
        spent = t.budget - t.remaining_budget
        left = max(0, t.max_players - purchased)
        
        # Calculate max bid
        base_price = tournament.default_base_price if (tournament and tournament.default_base_price is not None) else 1000.0
        if left <= 1:
            max_bid = t.remaining_budget
        else:
            max_bid = t.remaining_budget - (left - 1) * base_price
        max_bid = max(0, max_bid)

        # Get highest paid player
        highest_p = None
        if players:
            sold_players = [p for p in players if p.sold_price is not None]
            if sold_players:
                highest_p = max(sold_players, key=lambda p: p.sold_price)
                highest_paid_str = f"🏆 {highest_p.name} ₹{highest_p.sold_price:,.0f}"
            else:
                highest_paid_str = "-"
        else:
            highest_paid_str = "-"

        # 2. Stats Rows (Rows 2 to 5)
        stats = [
            f"Spent ₹{spent:,.0f} | Remaining ₹{t.remaining_budget:,.0f}",
            f"Purchased {purchased} | Left {left}",
            f"Max Bid ₹{max_bid:,.0f}",
            highest_paid_str
        ]

        for r_offset, stat_val in enumerate(stats, start=2):
            ws.merge_cells(start_row=r_offset, start_column=col_start, end_row=r_offset, end_column=col_end)
            c = ws.cell(row=r_offset, column=col_start)
            c.value = stat_val
            c.font = font_stats
            c.fill = fill_stats
            c.alignment = align_center
            
            # Format fill on the merged partner cell
            ws.cell(row=r_offset, column=col_end).fill = fill_stats

        # 3. Headers (Row 6)
        c_player = ws.cell(row=6, column=col_start)
        c_player.value = "Player"
        c_player.font = font_header
        c_player.fill = fill_subheader
        c_player.alignment = align_center
        c_player.border = border_thin

        c_bid = ws.cell(row=6, column=col_end)
        c_bid.value = "Bid"
        c_bid.font = font_header
        c_bid.fill = fill_subheader
        c_bid.alignment = align_center
        c_bid.border = border_thin

        # 4. Players List (Row 7 onwards)
        for p_idx, p in enumerate(players):
            r = 7 + p_idx
            
            # Gold highlight for the highest paid player
            if highest_p and p.id == highest_p.id:
                row_fill = PatternFill(start_color='FDE047', end_color='FDE047', fill_type='solid')
            else:
                row_fill = fill_even_row if p_idx % 2 == 0 else fill_odd_row

            c_pname = ws.cell(row=r, column=col_start)
            c_pname.value = f"{p_idx + 1}. {p.name}"
            c_pname.font = font_data
            c_pname.alignment = align_center
            c_pname.fill = row_fill
            c_pname.border = border_thin

            c_pbid = ws.cell(row=r, column=col_end)
            c_pbid.value = p.sold_price
            c_pbid.font = font_price
            c_pbid.alignment = align_center
            c_pbid.fill = row_fill
            c_pbid.border = border_thin
            c_pbid.number_format = '₹#,##0'

    # Set column widths nicely
    for col_idx in range(1, (len(teams) * 2) + 1):
        col_letter = get_column_letter(col_idx)
        # Alternate Player and Bid column widths
        if col_idx % 2 == 1:
            ws.column_dimensions[col_letter].width = 24
        else:
            ws.column_dimensions[col_letter].width = 14

    # Set row heights for premium spacing
    ws.row_dimensions[1].height = 28
    for r in range(2, 6):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[6].height = 22

    # Set borders for merged header cells
    for idx in range(len(teams)):
        col_start = (idx * 2) + 1
        col_end = (idx * 2) + 2
        for r in range(1, 6):
            ws.cell(row=r, column=col_start).border = border_thin
            ws.cell(row=r, column=col_end).border = border_thin

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Custom filename if exporting single team
    team_suffix = f"_{teams[0].name.replace(' ', '_')}" if len(teams) == 1 else "_teams"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{tournament.name}{team_suffix}.xlsx')


def _export_auction_excel(sold_players, tournament):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Auction Report'
    ws.append(['S.No', 'Player', 'Village', 'Playing Style', 'Base Price', 'Sold Price', 'Sold Team', 'Sold At'])
    for idx, p in enumerate(sold_players):
        ws.append([idx + 1, p.name, p.village, p.playing_style, p.base_price, p.sold_price,
                   p.sold_team.name if p.sold_team else '', str(p.sold_at) if p.sold_at else ''])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{tournament.name}_auction.xlsx')


# ---- PDF Exports ----

def _export_players_pdf(players, tournament):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(f'{tournament.name} - Players Report', styles['Title'])]

    data = [['S.No', 'Name', 'Village', 'Style', 'Base Price', 'Status', 'Sold Team', 'Sold Price']]
    for idx, p in enumerate(players):
        data.append([str(idx + 1), p.name, p.village or '', p.playing_style or '', f'₹{p.base_price}', p.status,
                     p.sold_team.name if p.sold_team else '', f'₹{p.sold_price}' if p.sold_price else ''])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0a1628')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True,
                     download_name=f'{tournament.name}_players.pdf')


def _export_teams_pdf(teams, tournament):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    styles = getSampleStyleSheet()
    
    title_style = styles['Title']
    title_style.fontSize = 18
    title_style.textColor = colors.HexColor('#0F172A')
    
    elements = [
        Paragraph(f'{tournament.name} - Teams & Squads Details', title_style),
        Spacer(1, 15)
    ]

    team_chunks = [teams[i:i + 4] for i in range(0, len(teams), 4)]

    for chunk in team_chunks:
        num_cols = len(chunk) * 2
        chunk_teams_players = [t.players.all() for t in chunk]
        max_players = max([len(p_list) for p_list in chunk_teams_players]) if chunk_teams_players else 0
        total_rows = 6 + max_players
        
        grid = [["" for _ in range(num_cols)] for _ in range(total_rows)]
        
        tbl_styles = [
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]
        
        for t_idx, t in enumerate(chunk):
            col_start = t_idx * 2
            col_end = t_idx * 2 + 1
            
            for r in range(5):
                tbl_styles.append(('SPAN', (col_start, r), (col_end, r)))
                
            grid[0][col_start] = t.name
            tbl_styles.extend([
                ('BACKGROUND', (col_start, 0), (col_end, 0), colors.HexColor('#0F172A')),
                ('TEXTCOLOR', (col_start, 0), (col_end, 0), colors.white),
                ('FONTNAME', (col_start, 0), (col_end, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (col_start, 0), (col_end, 0), 9),
            ])
            
            players = chunk_teams_players[t_idx]
            purchased = len(players)
            spent = t.budget - t.remaining_budget
            left = max(0, t.max_players - purchased)
            
            base_price = tournament.default_base_price if (tournament and tournament.default_base_price is not None) else 1000.0
            if left <= 1:
                max_bid = t.remaining_budget
            else:
                max_bid = t.remaining_budget - (left - 1) * base_price
            max_bid = max(0, max_bid)

            highest_p = None
            if players:
                sold_players = [p for p in players if p.sold_price is not None]
                if sold_players:
                    highest_p = max(sold_players, key=lambda p: p.sold_price)
                    highest_paid_str = f"Highest: {highest_p.name} Rs. {highest_p.sold_price:,.0f}"
                else:
                    highest_paid_str = "Highest: -"
            else:
                highest_paid_str = "Highest: -"
                
            grid[1][col_start] = f"Spent Rs. {spent:,.0f} | Rem Rs. {t.remaining_budget:,.0f}"
            grid[2][col_start] = f"Purchased {purchased} | Left {left}"
            grid[3][col_start] = f"Max Bid Rs. {max_bid:,.0f}"
            grid[4][col_start] = highest_paid_str
            
            for r in range(1, 5):
                tbl_styles.extend([
                    ('BACKGROUND', (col_start, r), (col_end, r), colors.HexColor('#F1F5F9')),
                    ('TEXTCOLOR', (col_start, r), (col_end, r), colors.HexColor('#1E293B')),
                    ('FONTNAME', (col_start, r), (col_end, r), 'Helvetica'),
                ])
                
            grid[5][col_start] = "Player"
            grid[5][col_end] = "Bid"
            tbl_styles.extend([
                ('BACKGROUND', (col_start, 5), (col_end, 5), colors.HexColor('#FEF08A')),
                ('TEXTCOLOR', (col_start, 5), (col_end, 5), colors.HexColor('#0F172A')),
                ('FONTNAME', (col_start, 5), (col_end, 5), 'Helvetica-Bold'),
            ])
            
            for p_idx, p in enumerate(players):
                r = 6 + p_idx
                grid[r][col_start] = f"{p_idx + 1}. {p.name}"
                grid[r][col_end] = f"Rs. {p.sold_price:,.0f}" if p.sold_price is not None else "-"
                
                if highest_p and p.id == highest_p.id:
                    row_bg = colors.HexColor('#FDE047')
                else:
                    row_bg = colors.HexColor('#F8FAFC') if p_idx % 2 == 0 else colors.white

                tbl_styles.extend([
                    ('BACKGROUND', (col_start, r), (col_end, r), row_bg),
                    ('TEXTCOLOR', (col_start, r), (col_end, r), colors.HexColor('#0F172A')),
                    ('FONTNAME', (col_start, r), (col_end, r), 'Helvetica'),
                ])
                
            for r in range(6 + len(players), total_rows):
                row_bg = colors.HexColor('#F8FAFC') if (r - 6) % 2 == 0 else colors.white
                tbl_styles.extend([
                    ('BACKGROUND', (col_start, r), (col_end, r), row_bg),
                ])

        col_widths = []
        for _ in range(len(chunk)):
            col_widths.extend([125, 65])
            
        t_table = Table(grid, colWidths=col_widths)
        t_table.setStyle(TableStyle(tbl_styles))
        elements.append(t_table)
        elements.append(Spacer(1, 25))

    doc.build(elements)
    output.seek(0)
    
    # Custom filename if exporting single team
    team_suffix = f"_{teams[0].name.replace(' ', '_')}" if len(teams) == 1 else "_teams"
    return send_file(output, mimetype='application/pdf', as_attachment=True,
                     download_name=f'{tournament.name}{team_suffix}.pdf')


def _export_auction_pdf(sold_players, tournament):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(f'{tournament.name} - Auction Report', styles['Title'])]

    data = [['S.No', 'Player', 'Village', 'Style', 'Base Price', 'Sold Price', 'Sold Team']]
    for idx, p in enumerate(sold_players):
        data.append([str(idx + 1), p.name, p.village or '', p.playing_style or '', f'₹{p.base_price}',
                     f'₹{p.sold_price}', p.sold_team.name if p.sold_team else ''])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0a1628')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True,
                     download_name=f'{tournament.name}_auction.pdf')
